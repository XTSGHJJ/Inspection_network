from openpyxl import load_workbook
from netmiko import ConnectHandler
from time import strftime,time
from  threading import Thread
from os import getcwd


def netmiko_dev_connect(ip,username,passwd,check_comm,dev_type,back_file,en_passwd=''):
    dev = {'device_type':dev_type,'ip':ip,'username':username,'password':passwd,'secret':en_passwd,'session_log': 'debug_test.log'}
    log_path = getcwd().replace('\\','/')
    try:
        connect = ConnectHandler(**dev)
        if not dev['secret'] == None: #判断是否需要enabale密码
            connect.enable()
        dev_name = connect.find_prompt().strip('<>') #匹配主机名
        now_time = strftime('-%Y-%m-%d-%H-%M-%S') #记录时间
        for comm in check_comm: #发送命令，保存输出
            #strip_prompt用于在输出的文本中会显示设备名
            #strip_command用于在输出的文本中显示输入的命令
            content = connect.send_command(comm,strip_prompt=False,strip_command=False) 
            with open(back_file + dev_name + now_time + '.txt','a',encoding='utf-8') as backup_conf: #以设备名+时间命名的txt输出文件
                backup_conf.write(content + '\n')
        print(ip+'------巡检完毕！')
        with open(log_path+'/log_file/success_device'+'.txt','a',encoding='utf-8') as success_file:
            success_file.write( str(ip) + '--成功执行完毕！' + '\n')
            
    except Exception as error_message:
        # print(str(error_message))
        with open(log_path+'/log_file/error_device'+'.txt','a',encoding='utf-8') as error_file:
            error_file.write( str(ip) + '-错误消息:' + str(error_message) + '\n')
            
def dev_judge(DeviceBrand):
    dev_value = DeviceBrand.lower().strip()
    pwd_dir=getcwd().replace('\\','/') #获取当前文件路径
    #Huawei
    if dev_value == 'huawei':
        hw_comm = pwd_dir + '/devInfo_file/hw_comm.txt'
        hw_bac_path = pwd_dir + '/backup_file/hw_bac_file/' #backup路径最后要加/，否则文件无法创建成功
        return ['huawei',hw_comm,hw_bac_path]
    #New H3C
    elif dev_value == 'h3c':
        h3c_comm = pwd_dir + '/devInfo_file/h3c_comm.txt'
        h3c_bac_path = pwd_dir + '/backup_file/h3c_bac_file/'
        return ['hp_comware',h3c_comm,h3c_bac_path]
    #Fortinet
    elif dev_value == 'forigate':
        ft_comm = pwd_dir + '/devInfo_file/ft_comm.txt'
        ft_bac_path = pwd_dir + '/backup_file/ft_bac_file/'
        return ['fortinet',ft_comm,ft_bac_path]
    #Cisco
    elif dev_value == 'cisco':
        cisco_comm = pwd_dir + '/devInfo_file/cisco_comm.txt'
        cisco_bac_path = pwd_dir + '/backup_file/cisco_bac_file/'
        return ['cisco_ios',cisco_comm,cisco_bac_path]
    #Hillstone
    elif dev_value == 'hillstone':
        sg_comm = pwd_dir + '/devInfo_file/sg_comm.txt'
        sg_bac_path = pwd_dir + '/backup_file/sg_bac_file/'
        return ['SG',sg_comm,sg_bac_path]
    else: 
        return 'None' #如果没判断出来返回一个none

def main():
    start_time = time() #记录时间
    wb = load_workbook('device_info.xlsx')
    ws =wb.active
    row_range = ws.max_row #行数
    therad_list = []
    for row_list in range(2,row_range+1):
        dev_info = ws[row_list]
        dev_temp = []
        for i in dev_info: #遍历后添加到列表
            dev_temp.append(i.value)
        if dev_temp[7] == None: #如果没有IP地址直接跳过
            continue
                
#----列表索引对应的值
        # device_area = dev_temp[0]        
        # device_name = dev_temp[1]
        # device_type(netmiko-type) = dev_temp[2]
        # device_type(switch or router) = dev_temp[3]
        # device_connect(ssh or telnet) = dev_temp[4]
        # device_user = dev_temp[5]
        # device_pwd = dev_temp[6]
        # device_ip = dev_temp[7]
        # device_enable-password = dev_temp[8]
#------------------------------------------------------
        
        judge_result = dev_judge(dev_temp[2]) #判断设备厂商类型        
        if not judge_result == 'None':
            if dev_temp[4].lower().strip() == 'telnet':
                judge_result[0]=judge_result[0]+'_telnet'

            check_command = open(judge_result[1],'r')
            thread_netmiko = Thread(target=netmiko_dev_connect,
            args=(dev_temp[7],
                  dev_temp[5],
                  dev_temp[6],
                  check_command,
                  judge_result[0],
                  judge_result[2],
                  dev_temp[8])
                  )
            thread_netmiko.start()
            therad_list.append(thread_netmiko)
        else:
            print('不支持此设备:',dev_temp[1])
    for y in therad_list:
        y.join()
        
    end_time = time()
    total_time = end_time - start_time
    print('共用时:'+ "%.2f" %total_time + '秒') 

if __name__ == '__main__':
    main()